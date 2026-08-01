# -*- coding: utf-8 -*-
"""
Swift POS — Backend API v1
Frappe v15 + ERPNext

All endpoints in this file are thin wrappers around native ERPNext DocTypes.
No business data is duplicated. Configuration is always resolved from the
`Swift POS Settings` single doctype (never hardcoded).

URL exposure:
Frappe whitelisted methods are natively callable at:
    /api/method/swift_pos.api.v1.api.<function_name>
To expose them under the cleaner /api/v1/... path used in the plan, add the
matching rewrite rules to hooks.py (see hooks_snippet.py in this folder).
"""

import unicodedata

import frappe
from frappe import _
from frappe.utils import now_datetime, add_to_date, cint, flt, date_diff, nowdate


# Returns are accepted only within this many days of the invoice posting date.
RETURN_WINDOW_DAYS = 5


# ---------------------------------------------------------------------------
# Section 0 — Centralized configuration helpers
# ---------------------------------------------------------------------------

def get_settings():
	"""Single source of truth for all configurable values."""
	return frappe.get_cached_doc("Swift POS Settings")


def get_pos_profile_doc():
	settings = get_settings()
	if not settings.default_pos_profile:
		frappe.throw(_("Swift POS Settings: default_pos_profile is not configured."))
	return frappe.get_cached_doc("POS Profile", settings.default_pos_profile)


def resolve_config():
	"""Returns fully resolved config dict used across the API and by the
	frontend bootstrap call (7.1)."""
	settings = get_settings()
	profile = get_pos_profile_doc()

	payment_modes = [p.mode_of_payment for p in profile.payments] if profile.payments else []

	return {
		"company": settings.default_company,
		"pos_profile": settings.default_pos_profile,
		"price_list": settings.default_price_list,
		"warehouse": profile.warehouse,
		"currency": profile.currency,
		"cost_center": getattr(profile, "cost_center", None),
		"payment_modes": payment_modes,
	}


def require_role(role):
	"""Raise 403 if the current user does not have the given role."""
	if role not in frappe.get_roles(frappe.session.user):
		frappe.throw(_("You are not permitted to perform this action."), frappe.PermissionError)


def _require_any_role(*roles):
	"""Raise 403 unless the user holds at least one of the given roles."""
	user_roles = frappe.get_roles(frappe.session.user)
	if not any(role in user_roles for role in roles):
		frappe.throw(_("You are not permitted to perform this action."), frappe.PermissionError)


# Fields the Storekeeper may edit through update_item. Kept next to the other
# access-control helpers so the allowlist is easy to audit.
EDITABLE_ITEM_FIELDS = ("item_name", "item_group", "description", "disabled")


def current_device_id():
	"""Frontend must send this header on every request."""
	device_id = frappe.get_request_header("X-Device-Id")
	if not device_id:
		frappe.throw(_("Missing X-Device-Id header."))
	return device_id


# ---------------------------------------------------------------------------
# Section 1 — Authentication
# ---------------------------------------------------------------------------

@frappe.whitelist(allow_guest=True, methods=["POST"])
def login(email=None, password=None):
	if not email or not password:
		frappe.throw(_("Email and password are required."))

	login_manager = frappe.auth.LoginManager()
	try:
		login_manager.authenticate(user=email, pwd=password)
		login_manager.post_login()
	except frappe.exceptions.AuthenticationError:
		frappe.local.response.http_status_code = 401
		frappe.throw(_("Invalid credentials."), frappe.AuthenticationError)

	roles = frappe.get_roles(frappe.session.user)
	role = "Swift Cashier" if "Swift Cashier" in roles else (
		"Swift Storekeeper" if "Swift Storekeeper" in roles else None
	)
	if not role:
		frappe.throw(_("User has no Swift role assigned (Cashier/Storekeeper)."))

	return {
		"user": frappe.session.user,
		"role": role,
		"full_name": frappe.get_cached_value("User", frappe.session.user, "full_name"),
		"sid": frappe.session.sid,
	}


@frappe.whitelist(methods=["POST"])
def logout():
	frappe.local.login_manager.logout()
	frappe.db.commit()
	return {"success": True}


@frappe.whitelist(methods=["GET"])
def me():
	roles = frappe.get_roles(frappe.session.user)
	role = "Swift Cashier" if "Swift Cashier" in roles else (
		"Swift Storekeeper" if "Swift Storekeeper" in roles else None
	)
	return {
		"user": frappe.session.user,
		"role": role,
		"full_name": frappe.get_cached_value("User", frappe.session.user, "full_name"),
	}


# ---------------------------------------------------------------------------
# Section 2 — POS Session Management (Cashier)
# ---------------------------------------------------------------------------

def _get_open_session_for_user(user):
	name = frappe.db.get_value(
		"POS Opening Entry",
		{"user": user, "status": "Open", "docstatus": 1},
		"name",
	)
	if not name:
		return None
	return frappe.get_doc("POS Opening Entry", name)


@frappe.whitelist(methods=["GET"])
def session_current():
	require_role("Swift Cashier")
	doc = _get_open_session_for_user(frappe.session.user)
	if not doc:
		return {"exists": False}

	opening_amount = 0
	if doc.balance_details:
		opening_amount = doc.balance_details[0].opening_amount

	return {
		"exists": True,
		"opening_entry": doc.name,
		"opening_time": doc.period_start_date,
		"opening_amount": opening_amount,
	}


@frappe.whitelist(methods=["POST"])
def session_open(opening_amount=None):
	require_role("Swift Cashier")

	if opening_amount is None:
		frappe.throw(_("opening_amount is required."))
	opening_amount = flt(opening_amount)
	if opening_amount < 0:
		frappe.throw(_("opening_amount must be zero or positive."))

	# Reconnect guard — never create a duplicate session.
	existing = _get_open_session_for_user(frappe.session.user)
	if existing:
		return session_current()

	settings = get_settings()
	profile = get_pos_profile_doc()
	device_id = current_device_id()

	# Multi-device guard.
	if not settings.allow_multi_device_session:
		other_device_open = frappe.db.exists(
			"POS Opening Entry",
			{
				"user": frappe.session.user,
				"status": "Open",
				"docstatus": 1,
				"custom_device_id": ["!=", device_id],
			},
		)
		if other_device_open:
			frappe.local.response.http_status_code = 409
			frappe.throw(_("An active session already exists on another device."))

	default_mode = profile.payments[0].mode_of_payment if profile.payments else None
	if not default_mode:
		frappe.throw(_("POS Profile has no Mode of Payment configured."))

	doc = frappe.new_doc("POS Opening Entry")
	doc.period_start_date = now_datetime()
	doc.posting_date = now_datetime().date()
	doc.company = settings.default_company
	doc.pos_profile = settings.default_pos_profile
	doc.user = frappe.session.user
	doc.set("balance_details", [])
	doc.append("balance_details", {
		"mode_of_payment": default_mode,
		"opening_amount": opening_amount,
	})
	# custom field to support device-scoped session lookups (see hooks/fixtures)
	doc.custom_device_id = device_id
	# Elevated: Swift Cashier is a custom role with no POS Opening Entry
	# permission. user is forced to the session user above, so a cashier cannot
	# open a session as anyone else. submit() re-checks, hence the flag.
	doc.flags.ignore_permissions = True
	doc.insert(ignore_permissions=True)
	doc.submit()

	return {
		"opening_entry": doc.name,
		"period_start_date": doc.period_start_date,
		"status": "Open",
	}


@frappe.whitelist(methods=["POST"])
def session_heartbeat(opening_entry=None, state="idle"):
	require_role("Swift Cashier")
	if state not in ("cart_active", "payment_open", "idle"):
		frappe.throw(_("Invalid state value."))

	doc = _get_open_session_for_user(frappe.session.user)
	if not doc or (opening_entry and doc.name != opening_entry):
		frappe.throw(_("No matching active session."))

	frappe.db.set_value(
		"POS Opening Entry",
		doc.name,
		{"custom_last_heartbeat": now_datetime(), "custom_heartbeat_state": state},
		update_modified=False,
	)
	return {"acknowledged": True}


def _build_closing_from_opening(closing, opening_doc, closing_amount):
	"""Populate payment_reconciliation on a POS Closing Entry by querying
	submitted Sales Invoices linked to the opening entry.
	Also subtracts any POS expenses (Journal Entries) made during this shift.

	pos_transactions is deliberately left empty. Its child doctype
	(POS Invoice Reference.pos_invoice) is a Link to POS Invoice and cannot hold a
	Sales Invoice name, and POS Closing Entry.on_submit feeds those rows to
	consolidate_pos_invoices(). Sales Invoices already posted their own stock, GL
	and payment on submit, so an empty table is what keeps closing from posting a
	second time. Closing remains a pure cash-reconciliation document."""

	opening_name = opening_doc.name

	# 1) Sales Invoices of this shift — the basis for cash reconciliation below.
	invoices = frappe.get_all(
		"Sales Invoice",
		filters={"custom_pos_opening_entry": opening_name, "docstatus": 1},
		fields=["name", "grand_total", "posting_date", "posting_time", "is_return"],
	)
	closing.set("pos_transactions", [])

	# 2) Build payment_reconciliation by aggregating all payment entries
	#    across every submitted Sales Invoice for this session.
	mode_totals = {}
	for inv in invoices:
		inv_doc = frappe.get_doc("Sales Invoice", inv.name)
		for p in inv_doc.payments:
			mode = p.mode_of_payment
			mode_totals.setdefault(mode, 0)
			mode_totals[mode] += flt(p.amount)

	# Also include modes from the opening entry balance_details that
	# may not have had any invoices (e.g. Cash with 0 sales).
	for bd in opening_doc.balance_details:
		mode_totals.setdefault(bd.mode_of_payment, 0)

	# 3) Calculate total POS expenses during this shift.
	#    Journal Entries tagged with [POS:<opening_name>] in user_remark.
	#    Each expense credits the cash account, so we sum the credit amounts.
	total_expenses = 0
	expense_jes = frappe.get_all(
		"Journal Entry",
		filters={
			"docstatus": 1,
			"user_remark": ["like", f"%[POS:{opening_name}]%"],
		},
		fields=["name"],
	)
	for je in expense_jes:
		je_doc = frappe.get_doc("Journal Entry", je.name)
		for row in je_doc.accounts:
			if flt(row.credit_in_account_currency) > 0:
				total_expenses += flt(row.credit_in_account_currency)

	closing.set("payment_reconciliation", [])
	expected_amount = 0
	for mode, total in mode_totals.items():
		opening_amt = 0
		for bd in opening_doc.balance_details:
			if bd.mode_of_payment == mode:
				opening_amt = flt(bd.opening_amount)
				break

		row_expected = flt(opening_amt) + flt(total)
		# Subtract expenses from cash expected amount
		if mode == "Cash":
			row_expected -= flt(total_expenses)

		row_closing = 0
		if mode == "Cash":
			row_closing = flt(closing_amount)
			expected_amount = row_expected

		closing.append("payment_reconciliation", {
			"mode_of_payment": mode,
			"opening_amount": opening_amt,
			"expected_amount": row_expected,
			"closing_amount": row_closing,
			"difference": flt(row_closing) - flt(row_expected),
		})

	return {"expected_amount": expected_amount, "total_expenses": flt(total_expenses)}


@frappe.whitelist(methods=["POST"])
def session_close(closing_amount=None):
	require_role("Swift Cashier")

	if closing_amount is None:
		frappe.throw(_("closing_amount is required."))
	closing_amount = flt(closing_amount)

	doc = _get_open_session_for_user(frappe.session.user)
	if not doc:
		frappe.local.response.http_status_code = 409
		frappe.throw(_("No active session to close."))

	# Block close if any draft invoice exists under this session.
	draft_exists = frappe.db.exists(
		"Sales Invoice",
		{"custom_pos_opening_entry": doc.name, "docstatus": 0},
	)
	if draft_exists:
		frappe.throw(_("Draft invoices exist for this session — resolve them before closing."))

	closing = frappe.new_doc("POS Closing Entry")
	closing.pos_opening_entry = doc.name
	closing.period_end_time = now_datetime()
	closing.posting_date = now_datetime().date()
	closing.user = frappe.session.user
	closing.company = doc.company
	closing.pos_profile = doc.pos_profile

	closing_result = _build_closing_from_opening(closing, doc, closing_amount)
	expected_amount = closing_result["expected_amount"]
	total_expenses = closing_result["total_expenses"]

	# No negative-stock override. It existed because consolidation posted every
	# sale's stock at closing time; stock now posts per sale and create_invoice
	# refuses lines it cannot cover, so closing never touches the Stock Ledger.
	closing.flags.ignore_permissions = True
	closing.insert()
	closing.submit()

	frappe.db.set_value("POS Opening Entry", doc.name, "status", "Closed")

	return {
		"closing_entry": closing.name,
		"expected_amount": expected_amount,
		"total_expenses": total_expenses,
		"difference": flt(closing_amount) - flt(expected_amount),
		"status": "Submitted",
	}


# ---------------------------------------------------------------------------
# Section 3 — POS Selling / Invoice
# ---------------------------------------------------------------------------

def _available_qty(item_code, company):
	"""On-hand qty for an item across every non-group warehouse of the company.

	The POS Profile warehouse is only one of several stores, so checking it alone
	reported 0 for items that are physically in stock elsewhere. Selling is gated
	on this total, and _sale_warehouse below picks the warehouse the stock is
	actually in.
	"""
	rows = frappe.get_all(
		"Bin",
		filters={"item_code": item_code, "warehouse": ["in", _stock_warehouses(company)]},
		fields=["warehouse", "actual_qty"],
	)
	return sum(flt(row.actual_qty) for row in rows)


def _stock_warehouses(company):
	"""Non-group warehouses of the company — the only ones that can hold stock."""
	return [
		row.name
		for row in frappe.get_all(
			"Warehouse",
			filters={"company": company, "is_group": 0, "disabled": 0},
			fields=["name"],
		)
	]


def _sale_warehouse(item_code, qty, company, preferred=None):
	"""Warehouse to draw `qty` of `item_code` from.

	Prefers the configured POS warehouse when it holds enough, otherwise falls
	back to the warehouse with the most stock. Returns None when no single
	warehouse can cover the line.
	"""
	rows = frappe.get_all(
		"Bin",
		filters={"item_code": item_code, "warehouse": ["in", _stock_warehouses(company)]},
		fields=["warehouse", "actual_qty"],
		order_by="actual_qty desc",
	)
	if preferred:
		for row in rows:
			if row.warehouse == preferred and flt(row.actual_qty) >= qty:
				return preferred
	for row in rows:
		if flt(row.actual_qty) >= qty:
			return row.warehouse
	return None


@frappe.whitelist(methods=["GET"])
def item_by_barcode(barcode=None):
	roles = frappe.get_roles(frappe.session.user)
	if "Swift Cashier" not in roles and "Swift Storekeeper" not in roles:
		frappe.throw(_("You are not permitted to perform this action."), frappe.PermissionError)
	if not barcode:
		frappe.throw(_("barcode is required."))

	item_code = frappe.db.get_value("Item Barcode", {"barcode": barcode}, "parent")
	if not item_code:
		frappe.local.response.http_status_code = 404
		frappe.throw(_("Item not found for barcode {0}").format(barcode))

	item = frappe.db.get_value(
		"Item", item_code,
		["item_code", "item_name", "stock_uom", "disabled", "image"],
		as_dict=True,
	)
	if not item or item.disabled:
		frappe.local.response.http_status_code = 410
		frappe.throw(_("Item is disabled."))

	config = resolve_config()
	rate = frappe.db.get_value(
		"Item Price",
		{"item_code": item_code, "price_list": config["price_list"]},
		"price_list_rate",
	) or 0

	# Stock across every warehouse of the company, not just the POS Profile one —
	# an item held in another store used to report 0 here.
	stock_qty = _available_qty(item_code, config["company"])

	# Refuse at scan time so an out-of-stock item never reaches the cart.
	if stock_qty <= 0:
		frappe.local.response.http_status_code = 409
		frappe.throw(_("{0} is out of stock.").format(item.item_name))

	return {
		"item_code": item.item_code,
		"item_name": item.item_name,
		"rate": flt(rate),
		"uom": item.stock_uom,
		"stock_qty": flt(stock_qty),
		"image": item.image,
	}


@frappe.whitelist(methods=["GET"])
def item_search(q=None):
	roles = frappe.get_roles(frappe.session.user)
	if "Swift Cashier" not in roles and "Swift Storekeeper" not in roles:
		frappe.throw(_("You are not permitted to perform this action."), frappe.PermissionError)
	if not q or len(q) < 2:
		frappe.throw(_("Query must be at least 2 characters."))

	items = frappe.get_all(
		"Item",
		filters={"disabled": 0},
		or_filters=[["item_name", "like", f"%{q}%"], ["item_code", "like", f"%{q}%"]],
		fields=["item_code", "item_name", "stock_uom", "image"],
		limit=20,
	)

	config = resolve_config()
	in_stock = []
	for item in items:
		# stock_qty was never returned here before, so the grid rendered "undefined".
		item["stock_qty"] = _available_qty(item.item_code, config["company"])
		if item["stock_qty"] <= 0:
			continue  # out of stock — not sellable, so not offered
		item["rate"] = flt(frappe.db.get_value(
			"Item Price",
			{"item_code": item.item_code, "price_list": config["price_list"]},
			"price_list_rate",
		) or 0)
		in_stock.append(item)

	return in_stock


@frappe.whitelist(methods=["POST"])
def create_invoice(items=None, payments=None, customer=None):
	require_role("Swift Cashier")

	if isinstance(items, str):
		items = frappe.parse_json(items)
	if isinstance(payments, str):
		payments = frappe.parse_json(payments)

	if not items:
		frappe.throw(_("items array cannot be empty."))
	if not payments:
		frappe.throw(_("payments array cannot be empty."))

	session = _get_open_session_for_user(frappe.session.user)
	if not session:
		frappe.local.response.http_status_code = 409
		frappe.throw(_("No active POS session — open a shift first."))

	config = resolve_config()

	# Sales Invoice, not POS Invoice: POSInvoice.on_submit omits update_stock_ledger()
	# and make_gl_entries() entirely, so stock and accounting only posted at closing
	# via consolidation. SalesInvoice.on_submit posts both immediately. is_pos=1 keeps
	# the payments table and POS behaviour; update_stock=1 drives the Stock Ledger.
	inv = frappe.new_doc("Sales Invoice")
	inv.is_pos = 1
	inv.pos_profile = config["pos_profile"]
	inv.company = config["company"]
	inv.customer = customer or frappe.db.get_value("POS Profile", config["pos_profile"], "customer")
	# Sales Invoice has no pos_opening_entry field; the shift link is a custom field
	# following the same convention as custom_device_id on POS Opening Entry.
	inv.custom_pos_opening_entry = session.name
	inv.set_warehouse = config["warehouse"]
	inv.update_stock = 1

	# No negative-stock override. Availability is enforced per line below, so the
	# sale is refused instead of driving a warehouse negative. The previous code
	# flipped Stock Settings.allow_negative_stock globally for the duration of the
	# sale, which affected every concurrent user and stayed on if the process died.

	# Same item can arrive on several cart lines; availability is checked against
	# the combined qty so two lines of 3 cannot together sell 5 units.
	requested = {}
	for row in items:
		if not row.get("item_code") or not row.get("qty"):
			frappe.throw(_("Each item requires item_code and qty."))
		if flt(row["qty"]) <= 0:
			frappe.throw(_("Quantity must be greater than zero."))
		requested[row["item_code"]] = requested.get(row["item_code"], 0) + flt(row["qty"])

	for row in items:
		item_exists = frappe.db.get_value(
			"Item", row["item_code"], ["disabled", "item_name"], as_dict=True
		)
		if not item_exists:
			frappe.local.response.http_status_code = 404
			frappe.throw(_("Item {0} not found.").format(row["item_code"]))
		if item_exists.get("disabled"):
			frappe.throw(_("Item {0} is disabled.").format(row["item_code"]))

		item_name = item_exists.get("item_name") or row["item_code"]
		total_qty = requested[row["item_code"]]
		available = _available_qty(row["item_code"], config["company"])

		if available <= 0:
			frappe.local.response.http_status_code = 409
			frappe.throw(_("{0} is out of stock.").format(item_name))
		if total_qty > available:
			frappe.local.response.http_status_code = 409
			frappe.throw(
				_("Only {0} of {1} available in stock.").format(flt(available), item_name)
			)

		# Draw from a warehouse that actually holds the stock rather than always
		# using the POS Profile warehouse.
		warehouse = _sale_warehouse(
			row["item_code"], total_qty, config["company"], config["warehouse"]
		)
		if not warehouse:
			frappe.local.response.http_status_code = 409
			frappe.throw(
				_(
					"{0} has {1} in stock but not in a single warehouse. Transfer stock before selling."
				).format(item_name, flt(available))
			)

		inv.append("items", {
			"item_code": row["item_code"],
			"qty": flt(row["qty"]),
			"rate": flt(row.get("rate") or 0) or None,
			"warehouse": warehouse,
		})

	inv.set_missing_values()
	inv.calculate_taxes_and_totals()

	grand_total = flt(inv.grand_total)

	# Validate payment modes exist; amounts will be overridden to grand_total below.
	for p in payments:
		if not p.get("mode_of_payment"):
			frappe.throw(_("Each payment requires mode_of_payment."))

	# Use a single payment row set to the actual grand_total (post-tax).
	# The frontend sends the pre-tax cart total which may differ after ERPNext applies taxes.
	mode = payments[0]["mode_of_payment"]
	inv.append("payments", {
		"mode_of_payment": mode,
		"amount": grand_total,
	})

	# Elevated: Swift Cashier is a custom role with no Sales Invoice permission.
	# Customer, warehouse, prices and totals are all resolved server-side above.
	inv.flags.ignore_permissions = True
	inv.insert(ignore_permissions=True)
	inv.submit()

	taxes = []
	for t in inv.taxes:
		taxes.append({
			"description": t.description,
			"tax_amount": flt(t.tax_amount),
			"rate": flt(t.rate),
		})

	# Check for low/negative stock items to warn the cashier
	stock_warnings = []
	for row in items:
		stock_qty = flt(frappe.db.get_value(
			"Bin", {"item_code": row["item_code"], "warehouse": config["warehouse"]}, "actual_qty"
		) or 0)
		if stock_qty < 0:
			stock_warnings.append(f"{row['item_code']}: {stock_qty} in stock")

	return {
		"invoice": inv.name,
		"grand_total": grand_total,
		"net_total": flt(inv.net_total),
		"taxes": taxes,
		"status": inv.status,
		"stock_warnings": stock_warnings,
	}

@frappe.whitelist(methods=["POST"])
def create_return(invoice_name, items=None, reason=None):
	require_role("Swift Cashier")

	if isinstance(items, str):
		items = frappe.parse_json(items)

	# Re-checks the return policy at submit time rather than trusting whenever
	# the screen was loaded; throws if the invoice is not returnable.
	original = _returnable_invoice(invoice_name)

	from erpnext.controllers.sales_and_purchase_return import make_return_doc
	return_doc = make_return_doc("Sales Invoice", invoice_name)

	# Put the stock back where it was sold from. create_invoice resolves a real
	# leaf warehouse per line (_sale_warehouse, filtered to is_group=0), so the
	# original rows are the authoritative source. Keyed by row name rather than
	# item_code because the same item can appear on two lines drawn from two
	# different warehouses.
	#
	# Clearing set_warehouse is the part that actually fixes the reported error:
	# make_return_doc copies it from the original, where it holds the configured
	# POS warehouse (a group node). ERPNext pushes that value down onto every row
	# during validation, so it overwrites the per-row warehouses and the Stock
	# Ledger Entry is then rejected by block_transactions_against_group_warehouse.
	return_doc.set_warehouse = None
	sold_from = {row.name: row.warehouse for row in original.items}
	for row in return_doc.items:
		warehouse = sold_from.get(row.sales_invoice_item)
		if warehouse:
			row.warehouse = warehouse

	# make_return_doc already sets every row's qty to what is still returnable
	# (original qty minus prior returns, sales_and_purchase_return.py:532), so
	# these are the authoritative ceilings. A fully returned invoice yields all
	# zeroes, which is what makes the duplicate-return check below reliable.
	# Summed per item because the client selects by item, not by row.
	remaining = {}
	for row in return_doc.items:
		remaining[row.item_code] = remaining.get(row.item_code, 0) + abs(flt(row.qty))
	if not any(remaining.values()):
		frappe.throw(_("Invoice {0} has already been fully returned.").format(invoice_name))

	if items:
		# Partial return: keep only the requested rows, clamped to `remaining`.
		# The client is not trusted with quantities — an over-return would post
		# stock and GL that never existed on the original sale.
		requested = {row["item_code"]: flt(row["qty"]) for row in items}

		for item_code, qty in requested.items():
			if qty <= 0:
				continue
			allowed = remaining.get(item_code, 0)
			if qty > allowed:
				frappe.throw(
					_("Cannot return {0} of {1}. Only {2} remaining.").format(
						qty, item_code, allowed
					)
				)

		# Spread each requested quantity across that item's rows, in order, so a
		# line split across two rows (and therefore possibly two warehouses)
		# returns from the rows it was actually sold on.
		outstanding = dict(requested)
		filtered_items = []
		for row in return_doc.items:
			qty = outstanding.get(row.item_code)
			if not qty or qty <= 0:
				continue
			take = min(qty, abs(flt(row.qty)))
			if take <= 0:
				continue

			# make_return_doc fills serial_no with every serial still returnable
			# for this row (sales_and_purchase_return.py:575), sized to the full
			# remaining qty. When returning fewer, the list has to be trimmed to
			# match or ERPNext rejects the row for serial/qty mismatch.
			if row.get("serial_no"):
				serials = [s for s in str(row.serial_no).split("\n") if s.strip()]
				if len(serials) > take:
					row.serial_no = "\n".join(serials[: int(take)])

			row.qty = -take
			outstanding[row.item_code] = qty - take
			filtered_items.append(row)
		if not filtered_items:
			frappe.throw(_("None of the requested items match the original invoice."))
		return_doc.items = filtered_items

	if reason:
		# Native Sales Invoice field; no custom field needed for the return note.
		return_doc.remarks = reason

	# Elevated: same reason as create_invoice. The return is built from the original
	# invoice, and quantities are clamped to what that invoice actually contained.
	return_doc.flags.ignore_permissions = True
	return_doc.insert(ignore_permissions=True)
	return_doc.submit()

	return {"return_invoice": return_doc.name, "status": "Return"}


def _returnable_invoice(invoice_name):
	"""Load a Sales Invoice and enforce the return policy, or throw.

	Shared by get_invoice and create_return so the screen cannot show a return
	the submit would refuse, and so the window is re-checked at submit time
	rather than trusted from whenever the screen was loaded."""
	if not frappe.db.exists("Sales Invoice", invoice_name):
		frappe.local.response.http_status_code = 404
		frappe.throw(_("Invoice {0} not found.").format(invoice_name))

	doc = frappe.get_doc("Sales Invoice", invoice_name)

	if doc.docstatus == 0:
		frappe.throw(_("Invoice {0} is a draft and cannot be returned.").format(invoice_name))
	if doc.docstatus == 2:
		frappe.throw(_("Invoice {0} is cancelled and cannot be returned.").format(invoice_name))
	if doc.is_return:
		frappe.throw(_("{0} is itself a return and cannot be returned.").format(invoice_name))

	days = date_diff(nowdate(), doc.posting_date)
	if days > RETURN_WINDOW_DAYS:
		frappe.throw(
			_("Invoice {0} is {1} days old. Returns are only accepted within {2} days.").format(
				invoice_name, days, RETURN_WINDOW_DAYS
			)
		)

	return doc


@frappe.whitelist(methods=["GET"])
def get_invoice(invoice_name):
	require_role("Swift Cashier")

	# Deliberately not restricted to the caller's own session. A return is
	# presented days later, by whichever cashier is on shift, so the previous
	# ownership check rejected virtually every genuine return. The return policy
	# in _returnable_invoice is the control that matters here, and the response
	# is limited to the fields the Return screen renders.
	doc = _returnable_invoice(invoice_name)

	from erpnext.controllers.sales_and_purchase_return import make_return_doc
	returnable = make_return_doc("Sales Invoice", invoice_name)
	remaining = {row.item_code: abs(flt(row.qty)) for row in returnable.items}

	return {
		"name": doc.name,
		"customer": doc.customer,
		"customer_name": doc.customer_name,
		"posting_date": str(doc.posting_date),
		"posting_time": str(doc.posting_time),
		"currency": doc.currency,
		"net_total": flt(doc.net_total),
		"total_taxes_and_charges": flt(doc.total_taxes_and_charges),
		"discount_amount": flt(doc.discount_amount),
		"grand_total": flt(doc.grand_total),
		"items": [
			{
				"item_code": row.item_code,
				"item_name": row.item_name,
				"uom": row.uom,
				"rate": flt(row.rate),
				"amount": flt(row.amount),
				"discount_amount": flt(row.discount_amount),
				"qty_sold": flt(row.qty),
				"qty_returned": flt(row.qty) - remaining.get(row.item_code, 0),
				"remaining_qty": remaining.get(row.item_code, 0),
			}
			for row in doc.items
		],
	}


@frappe.whitelist(methods=["GET"])
def session_invoices(opening_entry):
	require_role("Swift Cashier")

	owner = frappe.db.get_value("POS Opening Entry", opening_entry, "user")
	if owner != frappe.session.user:
		frappe.local.response.http_status_code = 403
		frappe.throw(_("Not your session."))

	return frappe.get_all(
		"Sales Invoice",
		filters={"custom_pos_opening_entry": opening_entry, "docstatus": 1},
		fields=["name", "grand_total", "is_return", "posting_date", "posting_time"],
	)


# ---------------------------------------------------------------------------
# Section 4 — Storekeeper: Items & Barcodes
# ---------------------------------------------------------------------------

@frappe.whitelist(methods=["POST"])
def create_item(item_name=None, item_group=None, uom=None, opening_stock=0, warehouse=None, item_code=None, barcodes=None, valuation_rate=None, selling_price=None):
	require_role("Swift Storekeeper")

	if not item_name or not item_group or not uom:
		frappe.throw(_("item_name, item_group and uom are required."))

	if not frappe.db.exists("Item Group", item_group):
		frappe.throw(_("Item Group {0} not found.").format(item_group))

	config = resolve_config()

	naming_by = frappe.db.get_single_value("Stock Settings", "item_naming_by") or "Item Code"

	doc = frappe.new_doc("Item")
	doc.item_name = item_name
	doc.item_group = item_group
	doc.stock_uom = uom
	doc.is_stock_item = 1

	target_warehouse = warehouse or config["warehouse"]

	if item_code and item_code.strip():
		doc.item_code = item_code.strip()
	elif naming_by == "Item Code":
		doc.item_code = item_name
	if flt(opening_stock) > 0:
		doc.opening_stock = flt(opening_stock)
		doc.valuation_rate = flt(valuation_rate) if flt(valuation_rate) > 0 else 0
		doc.warehouse = target_warehouse

	# Set default warehouse in Item Defaults child table
	doc.append("item_defaults", {
		"company": config["company"],
		"default_warehouse": target_warehouse,
	})

	if barcodes:
		if isinstance(barcodes, str):
			barcodes = frappe.parse_json(barcodes)
		seen = set()
		for bc in barcodes:
			bc = str(bc).strip()
			if not bc or bc in seen:
				continue
			seen.add(bc)
			owner = _barcode_owner(bc)
			if owner:
				frappe.local.response.http_status_code = 409
				frappe.throw(_("Barcode already assigned to item {0}.").format(owner))
			doc.append("barcodes", {"barcode": bc})

	# Every item carries a barcode so it is scannable and exportable straight
	# away — generated only when none was supplied.
	_ensure_barcode(doc)

	# Elevated: create_item is gated on Swift Storekeeper, which holds no Item
	# permission in ERPNext, and every field is validated above. Without this the
	# manual "Add Item" screen fails exactly like the import did.
	doc.insert(ignore_permissions=True)

	# Create selling price in the configured price list
	if flt(selling_price) > 0:
		price_list = config.get("price_list") or frappe.db.get_single_value("Selling Settings", "selling_price_list")
		if price_list:
			item_price = frappe.new_doc("Item Price")
			item_price.item_code = doc.item_code
			item_price.price_list = price_list
			item_price.price_list_rate = flt(selling_price)
			item_price.insert(ignore_permissions=True)

	return {
		"item_code": doc.item_code,
		"name": doc.name,
		"barcode": doc.barcodes[0].barcode if doc.get("barcodes") else None,
	}


@frappe.whitelist(methods=["PUT"])
def update_item(item_code, **fields):
	"""Update a limited, explicit set of Item fields.

	Only EDITABLE_ITEM_FIELDS are accepted. Previously this accepted any field
	the Item doctype exposed, which let a Storekeeper set valuation_rate,
	disabled, is_stock_item and similar. Prices and stock have dedicated,
	validated endpoints (see update_inventory_item).
	"""
	require_role("Swift Storekeeper")
	fields.pop("cmd", None)

	if not frappe.db.exists("Item", item_code):
		frappe.local.response.http_status_code = 404
		frappe.throw(_("Item {0} not found.").format(item_code))

	rejected = [key for key in fields if key not in EDITABLE_ITEM_FIELDS]
	if rejected:
		frappe.throw(
			_("These fields cannot be edited here: {0}.").format(", ".join(sorted(rejected)))
		)

	doc = frappe.get_doc("Item", item_code)
	for key, value in fields.items():
		doc.set(key, value)
	# Elevated: gated on Swift Storekeeper, and EDITABLE_ITEM_FIELDS above rejects
	# anything outside the allowlist before this point.
	doc.save(ignore_permissions=True)

	return {
		"item_code": doc.name,
		"item_name": doc.item_name,
		"item_group": doc.item_group,
		"stock_uom": doc.stock_uom,
		"disabled": doc.disabled,
	}


@frappe.whitelist(methods=["GET"])
def get_item(item_code):
	_require_any_role("Swift Cashier", "Swift Storekeeper")

	if not frappe.db.exists("Item", item_code):
		frappe.local.response.http_status_code = 404
		frappe.throw(_("Item {0} not found.").format(item_code))

	# Explicit field list — as_dict() would leak every internal Item column.
	item = frappe.db.get_value(
		"Item",
		item_code,
		[
			"name as item_code",
			"item_name",
			"item_group",
			"stock_uom",
			"description",
			"image",
			"disabled",
			"is_stock_item",
			"has_serial_no",
			"has_batch_no",
			"valuation_rate",
		],
		as_dict=True,
	)
	item["barcodes"] = frappe.get_all(
		"Item Barcode", filters={"parent": item_code}, fields=["barcode", "barcode_type"]
	)
	return item


def _barcode_owner(barcode):
	return frappe.db.get_value("Item Barcode", {"barcode": barcode}, "parent")


@frappe.whitelist(methods=["GET"])
def validate_barcode(barcode):
	require_role("Swift Storekeeper")
	owner = _barcode_owner(barcode)
	if owner:
		return {"available": False, "assigned_to": owner}
	return {"available": True}


@frappe.whitelist(methods=["POST"])
def add_item_barcode(item_code, barcode):
	require_role("Swift Storekeeper")

	owner = _barcode_owner(barcode)
	if owner and owner != item_code:
		frappe.local.response.http_status_code = 409
		frappe.throw(_("Barcode already assigned to item {0}.").format(owner))

	doc = frappe.get_doc("Item", item_code)
	if not any(row.barcode == barcode for row in doc.barcodes):
		doc.append("barcodes", {"barcode": barcode})
		# Elevated: gated on Swift Storekeeper; only the barcodes grid is touched.
		doc.save(ignore_permissions=True)

	return {"success": True, "barcode": barcode}


@frappe.whitelist(methods=["DELETE"])
def remove_item_barcode(item_code, barcode):
	require_role("Swift Storekeeper")

	doc = frappe.get_doc("Item", item_code)
	rows = [row for row in doc.barcodes if row.barcode == barcode]
	if not rows:
		frappe.local.response.http_status_code = 404
		frappe.throw(_("Barcode not found on this item."))

	doc.barcodes = [row for row in doc.barcodes if row.barcode != barcode]
	# Elevated: same reason as add_item_barcode.
	doc.save(ignore_permissions=True)
	return {"success": True}


@frappe.whitelist(methods=["POST"])
def add_serial_number(item_code, serial_no):
	require_role("Swift Storekeeper")

	has_serial = frappe.db.get_value("Item", item_code, "has_serial_no")
	if not has_serial:
		frappe.throw(_("Item {0} is not configured for serial tracking.").format(item_code))

	serials = serial_no if isinstance(serial_no, list) else [serial_no]
	if isinstance(serial_no, str) and serial_no.startswith("["):
		serials = frappe.parse_json(serial_no)

	created = []
	for sn in serials:
		if frappe.db.exists("Serial No", sn):
			frappe.local.response.http_status_code = 409
			frappe.throw(_("Serial number {0} already exists.").format(sn))
		doc = frappe.new_doc("Serial No")
		doc.serial_no = sn
		doc.item_code = item_code
		# Elevated: endpoint is gated on Swift Storekeeper, which holds no Serial No
		# permission; both fields are validated above.
		doc.insert(ignore_permissions=True)
		created.append(doc.name)

	return {"created": created}


# ---------------------------------------------------------------------------
# Section 5 — Storekeeper: Stock Entry
# ---------------------------------------------------------------------------

@frappe.whitelist(methods=["POST"])
def create_stock_entry(stock_entry_type=None, items=None):
	require_role("Swift Storekeeper")

	if isinstance(items, str):
		items = frappe.parse_json(items)
	if stock_entry_type not in ("Material Receipt", "Material Transfer", "Material Issue"):
		frappe.throw(_("Invalid stock_entry_type."))
	if not items:
		frappe.throw(_("items array cannot be empty."))

	config = resolve_config()
	doc = frappe.new_doc("Stock Entry")
	doc.stock_entry_type = stock_entry_type
	doc.company = config["company"]

	for row in items:
		entry = {"item_code": row["item_code"], "qty": flt(row["qty"])}
		if stock_entry_type == "Material Receipt":
			entry["t_warehouse"] = row.get("t_warehouse") or config["warehouse"]
		elif stock_entry_type == "Material Issue":
			entry["s_warehouse"] = row.get("s_warehouse") or config["warehouse"]
		else:  # Material Transfer
			if not row.get("s_warehouse") or not row.get("t_warehouse"):
				frappe.throw(_("Material Transfer requires both s_warehouse and t_warehouse."))
			entry["s_warehouse"] = row["s_warehouse"]
			entry["t_warehouse"] = row["t_warehouse"]
		doc.append("items", entry)

	# The warehouses above are client-supplied, so they are validated against the
	# resolved company before the write is elevated. Without this check, raising
	# permissions would let a Storekeeper move stock into another company's
	# warehouse.
	for entry in doc.items:
		for field in ("s_warehouse", "t_warehouse"):
			name = entry.get(field)
			if not name:
				continue
			warehouse = frappe.db.get_value(
				"Warehouse", name, ["company", "is_group"], as_dict=True
			)
			if not warehouse:
				frappe.throw(_("Warehouse {0} does not exist.").format(name))
			if warehouse.is_group:
				frappe.throw(_("Warehouse {0} is a group and cannot hold stock.").format(name))
			if warehouse.company != config["company"]:
				frappe.throw(_("Warehouse {0} does not belong to {1}.").format(name, config["company"]))

	# Elevated: Swift Storekeeper holds no Stock Entry permission. The flag is set
	# on the document because submit() re-checks permissions independently.
	doc.flags.ignore_permissions = True
	doc.insert(ignore_permissions=True)
	doc.submit()

	return {"stock_entry": doc.name, "status": "Submitted"}


@frappe.whitelist(methods=["GET"])
def get_stock_entry(name):
	require_role("Swift Storekeeper")
	return frappe.get_doc("Stock Entry", name).as_dict()


# ---------------------------------------------------------------------------
# Section 6 — Read-only reference data
# ---------------------------------------------------------------------------

@frappe.whitelist(methods=["GET"])
def list_warehouses():
	require_role("Swift Storekeeper")
	return frappe.get_all("Warehouse", filters={"disabled": 0}, fields=["name", "warehouse_name"])


@frappe.whitelist(methods=["GET"])
def list_import_warehouses():
	"""Warehouses the Storekeeper can import stock into.

	Separate from `list_warehouses`, which other screens consume as a bare list.
	Group warehouses are excluded because they cannot hold stock, and the list is
	scoped to the resolved company so the import screen cannot offer a warehouse
	that `_import_config` would then reject.
	"""
	require_role("Swift Storekeeper")
	company = _resolve_company(None)

	filters = {"is_group": 0, "disabled": 0}
	if company:
		filters["company"] = company

	warehouses = frappe.get_all(
		"Warehouse",
		filters=filters,
		fields=["name", "warehouse_name"],
		order_by="warehouse_name asc",
	)

	# The screen preselects this one, matching what an import would use if the
	# Storekeeper does not choose.
	default = None
	try:
		default = _import_config(for_write=False)["warehouse"]
	except Exception:
		pass

	return {"warehouses": warehouses, "default": default}


@frappe.whitelist(methods=["GET"])
def list_item_groups():
	require_role("Swift Storekeeper")
	return frappe.get_all("Item Group", fields=["name", "parent_item_group", "is_group"])


@frappe.whitelist(methods=["GET"])
def list_suppliers():
	require_role("Swift Storekeeper")
	return frappe.get_all(
		"Supplier",
		filters={"disabled": 0},
		fields=["name", "supplier_name"],
		order_by="supplier_name asc",
	)


# ---------------------------------------------------------------------------
# Section 7 — Settings bootstrap
# ---------------------------------------------------------------------------

@frappe.whitelist(methods=["GET"])
def pos_config():
	# Both roles need this: the cashier for the expense accounts and the
	# storekeeper for the warehouse/company defaults.
	_require_any_role("Swift Cashier", "Swift Storekeeper")
	config = resolve_config()
	# Include expense accounts for the cashier expense feature
	company = config["company"]
	expense_accounts = frappe.get_all(
		"Account",
		filters={"company": company, "root_type": "Expense", "is_group": 0},
		fields=["name", "account_name"],
		order_by="account_name",
	)
	config["expense_accounts"] = expense_accounts
	return config


# ---------------------------------------------------------------------------
# Section 8 — Cashier Expenses (Payment Entry)
# ---------------------------------------------------------------------------

@frappe.whitelist(methods=["POST"])
def create_expense(amount=None, expense_account=None, remarks=None):
	require_role("Swift Cashier")

	amount = flt(amount)
	if amount <= 0:
		frappe.throw(_("Amount must be greater than zero."))
	if not expense_account:
		frappe.throw(_("Expense account is required."))

	# Get current POS session
	session = _get_open_session_for_user(frappe.session.user)
	if not session:
		frappe.local.response.http_status_code = 409
		frappe.throw(_("No active POS session — open a shift first."))

	config = resolve_config()
	company = config["company"]

	# The account must be a real expense leaf in the configured company. Checking
	# only that the account *exists* would let a cashier credit POS cash against
	# any account on the site (including other companies' balance-sheet accounts).
	account = frappe.db.get_value(
		"Account",
		expense_account,
		["company", "root_type", "is_group", "disabled"],
		as_dict=True,
	)
	if not account:
		frappe.throw(_("Account {0} not found.").format(expense_account))
	if account.company != company:
		frappe.throw(_("Account {0} does not belong to {1}.").format(expense_account, company))
	if account.is_group:
		frappe.throw(_("Account {0} is a group account.").format(expense_account))
	if account.disabled:
		frappe.throw(_("Account {0} is disabled.").format(expense_account))
	if account.root_type != "Expense":
		frappe.throw(_("Account {0} is not an expense account.").format(expense_account))

	# Get the POS cash account from Mode of Payment "Cash"
	cash_account = frappe.db.get_value(
		"Mode of Payment Account",
		{"parent": "Cash", "company": company},
		"default_account",
	)
	if not cash_account:
		frappe.throw(_("No cash account configured for POS. Check Mode of Payment 'Cash'."))

	expense_name = frappe.db.get_value("Account", expense_account, "account_name") or expense_account
	remark = remarks or f"POS Expense: {expense_name}"

	je = frappe.new_doc("Journal Entry")
	je.voucher_type = "Journal Entry"
	je.posting_date = frappe.utils.today()
	je.company = company
	# Tag with POS session so we can find expenses for this shift
	je.user_remark = f"[POS:{session.name}] {remark}"

	if config.get("cost_center"):
		cost_center = config["cost_center"]
	else:
		cost_center = None

	# Debit the expense account (increase expense)
	je.append("accounts", {
		"account": expense_account,
		"debit_in_account_currency": amount,
		"credit_in_account_currency": 0,
		"cost_center": cost_center,
	})

	# Credit the cash account (decrease cash)
	je.append("accounts", {
		"account": cash_account,
		"debit_in_account_currency": 0,
		"credit_in_account_currency": amount,
		"cost_center": cost_center,
	})

	# Elevated: Swift Cashier holds no Journal Entry permission. Accounts, company
	# and cost centre are all resolved and validated server-side above. The flag is
	# set on the document because submit() re-checks permissions independently.
	je.flags.ignore_permissions = True
	je.insert(ignore_permissions=True)
	je.submit()

	return {
		"payment_entry": je.name,
		"amount": amount,
		"expense_account": expense_account,
	}


# ---------------------------------------------------------------------------
# Section 9 — Auto-close scheduled job (registered via hooks.py scheduler_events)
# ---------------------------------------------------------------------------

def auto_close_inactive_sessions():
	"""Run every 5 minutes via cron (see hooks_snippet.py)."""
	settings = get_settings()
	if not settings.auto_close_enabled:
		return

	timeout_minutes = cint(settings.session_timeout_minutes) or 60
	threshold = add_to_date(now_datetime(), minutes=-timeout_minutes)

	open_sessions = frappe.get_all(
		"POS Opening Entry",
		filters={"status": "Open", "docstatus": 1},
		fields=["name", "user", "period_start_date", "custom_last_heartbeat", "custom_heartbeat_state"],
	)

	for session in open_sessions:
		last_activity = session.custom_last_heartbeat or session.period_start_date
		last_invoice_time = frappe.db.get_value(
			"Sales Invoice",
			{"custom_pos_opening_entry": session.name, "docstatus": 1},
			"modified",
			order_by="modified desc",
		)
		if last_invoice_time and last_invoice_time > last_activity:
			last_activity = last_invoice_time

		if last_activity > threshold:
			continue  # still active

		if session.custom_heartbeat_state in ("cart_active", "payment_open"):
			continue  # explicit exception rule

		draft_exists = frappe.db.exists(
			"Sales Invoice", {"custom_pos_opening_entry": session.name, "docstatus": 0}
		)
		if draft_exists:
			continue  # explicit exception rule

		_auto_close_session(session.name, session.user)


def _auto_close_session(opening_entry_name, user):
	frappe.set_user(user)
	try:
		doc = frappe.get_doc("POS Opening Entry", opening_entry_name)
		closing = frappe.new_doc("POS Closing Entry")
		closing.pos_opening_entry = doc.name
		closing.period_end_time = now_datetime()
		closing.posting_date = now_datetime().date()
		closing.user = user
		closing.company = doc.company
		closing.pos_profile = doc.pos_profile
		closing.custom_auto_closed = 1

		_build_closing_from_opening(closing, doc, 0)

		closing.flags.ignore_permissions = True
		closing.insert()
		closing.submit()
		frappe.db.set_value("POS Opening Entry", doc.name, "status", "Closed")
		frappe.db.commit()
	finally:
		frappe.set_user("Administrator")


# ---------------------------------------------------------------------------
# Section 10 — Storekeeper: Inventory Import / Export (Excel)
# ---------------------------------------------------------------------------
#
# Design notes (all verified against ERPNext v15 source before implementing):
#
# * Configuration comes from native ERPNext singles, never hardcoded:
#     - Stock Settings.item_group        -> default Item Group
#     - Stock Settings.stock_uom         -> default Stock UOM
#     - Stock Settings.default_warehouse -> falls back to POS Profile.warehouse
#     - Buying Settings.buying_price_list -> buying Item Price target
#     - Buying Settings.supplier_group   -> group for auto-created Suppliers
#     - Swift POS Settings.default_price_list -> selling Item Price target
# * Barcodes are stored with an EMPTY barcode_type on purpose. Item.validate_barcode()
#   runs a check-digit test via the `barcodenumber` package for known types
#   (EAN/UPC-A/...), and a randomly generated 12-digit number will almost never
#   satisfy a UPC-A checksum. An empty type skips that validation while the
#   `Item Barcode.barcode` column stays UNIQUE at the DB level.
# * Quantity is applied as an ABSOLUTE target via Stock Reconciliation, so
#   re-importing the same sheet is idempotent. ERPNext fills expense_account and
#   cost_center from Company defaults itself, and drops rows whose qty+rate are
#   unchanged — if every row is unchanged it raises
#   EmptyStockReconciliationItemsError, which we treat as "nothing to do".
# * Item Price rows are updated in place; Item Price.check_duplicates() forbids
#   a second row for the same item/price-list/qty combination.

IMPORT_COLUMN_ALIASES = {
	"name": "item_name",
	"item": "item_name",
	"item name": "item_name",
	"item_name": "item_name",
	"description": "item_name",
	"qty": "qty",
	"quantity": "qty",
	"stock": "qty",
	"current stock": "qty",
	"supplier": "supplier",
	"vendor": "supplier",
	"cost price": "cost_price",
	"cost": "cost_price",
	"buying price": "cost_price",
	"buying_price": "cost_price",
	"purchase price": "cost_price",
	"selling price": "selling_price",
	"selling_price": "selling_price",
	"price": "selling_price",
	"barcode": "barcode",
}

REQUIRED_IMPORT_COLUMNS = ("item_name", "qty")

# Export column order. Labels are the untranslated English header text on
# purpose: they must round-trip through IMPORT_COLUMN_ALIASES so an exported
# sheet can be edited and re-imported. Translating them at module load time
# would also freeze them to whichever language was active on first import.
EXPORT_COLUMNS = (
	("item_name", "Name"),
	("barcode", "Barcode"),
	("qty", "QTY"),
	("supplier", "Supplier"),
	("cost_price", "Cost Price"),
	("selling_price", "Selling Price"),
)


# Invisible characters Excel and Windows keyboards leave inside Arabic cells.
# Built from ordinals on purpose: as literals they are unreviewable in a diff and
# indistinguishable from each other in an editor.
_INVISIBLE_CHARS = tuple(
	chr(code)
	for code in (
		0x200B,  # zero width space
		0x200C,  # zero width non-joiner
		0x200D,  # zero width joiner
		0x200E,  # left-to-right mark
		0x200F,  # right-to-left mark
		0x061C,  # arabic letter mark
		0x202A,  # left-to-right embedding
		0x202B,  # right-to-left embedding
		0x202C,  # pop directional formatting
		0x202D,  # left-to-right override
		0x202E,  # right-to-left override
		0x2066,  # left-to-right isolate
		0x2067,  # right-to-left isolate
		0x2068,  # first strong isolate
		0x2069,  # pop directional isolate
		0xFEFF,  # byte order mark / zero width no-break space
		0x00AD,  # soft hyphen
	)
)


def _normalize_text(value):
	"""Normalize a spreadsheet cell to comparable, storable text.

	Arabic sheets exported from Excel routinely carry characters that are
	invisible but not equal to a space: NBSP (U+00A0), narrow NBSP (U+202F),
	Arabic tatweel padding, and the bidi marks Excel inserts around mixed
	Arabic/Latin runs. Two rows that look identical then produce two different
	keys, so the same item is created twice; the second insert collides on
	Item.name and the row vanishes from the import.

	NFC composition is applied because Arabic letters with diacritics have both a
	precomposed and a decomposed encoding, and MariaDB compares the byte forms.
	NFC is the composing form, so the text is never rewritten into a different
	script or stripped of its diacritics: Arabic is preserved exactly, only its
	encoding is made canonical.
	"""
	if value is None:
		return ""

	text = value if isinstance(value, str) else str(value)

	# Bidi and joiner controls carry no meaning for identity comparison, and Excel
	# adds them unpredictably. Removed before anything else so they cannot survive
	# as part of a "word". Written as escapes so this source stays reviewable.
	for invisible in _INVISIBLE_CHARS:
		text = text.replace(invisible, "")

	# NFC first, then split()/join collapses every Unicode space class -- NBSP,
	# narrow NBSP, U+2000..U+200A, ideographic space, tabs and newlines from
	# multi-line cells -- into single plain spaces, and trims the ends.
	text = unicodedata.normalize("NFC", text)

	return " ".join(text.split())


def _match_key(value):
	"""Comparison key for item and supplier names.

	Normalized text, case-folded for the Latin parts. Arabic has no case, so
	casefold is a no-op there; it only prevents "Battery 48V" and "battery 48v"
	from being treated as two different items.
	"""
	return _normalize_text(value).casefold()


def _first_existing(doctype, filters=None, order_by="creation asc"):
	"""Return the name of the first matching record, or None.

	Used to auto-discover a usable default when no setting names one. Ordered by
	creation so the choice is stable across calls rather than arbitrary.
	"""
	found = frappe.get_all(
		doctype, filters=filters or {}, pluck="name", order_by=order_by, limit=1
	)
	return found[0] if found else None


def _resolve_company(preferred=None):
	"""Configured company, else the only/first existing one."""
	if preferred and frappe.db.exists("Company", preferred):
		return preferred

	global_default = frappe.defaults.get_global_default("company")
	if global_default and frappe.db.exists("Company", global_default):
		return global_default

	return _first_existing("Company")


def _resolve_item_group(preferred=None):
	"""Configured item group, else a sane discovered one.

	Prefers a non-group leaf so Items attach to a real node; ERPNext ships
	"Products"/"All Item Groups" on a fresh site, and only "All Item Groups" is
	a container.
	"""
	if preferred and frappe.db.exists("Item Group", preferred):
		return preferred

	# A leaf group is the correct parent for an Item.
	leaf = _first_existing("Item Group", {"is_group": 0})
	if leaf:
		return leaf

	return _first_existing("Item Group")


def _resolve_stock_uom(preferred=None):
	"""Configured stock UOM, else "Nos" (ERPNext default), else any UOM."""
	if preferred and frappe.db.exists("UOM", preferred):
		return preferred

	for candidate in ("Nos", "Unit", "Each"):
		if frappe.db.exists("UOM", candidate):
			return candidate

	return _first_existing("UOM")


def _resolve_warehouse(company, preferred=None):
	"""Configured warehouse, else a discovered non-group warehouse.

	Stock can only post to a leaf warehouse, so group nodes are excluded. Scoped
	to the resolved company first, then relaxed if that yields nothing.
	"""
	if preferred and frappe.db.exists("Warehouse", preferred):
		return preferred

	if company:
		scoped = _first_existing("Warehouse", {"is_group": 0, "company": company, "disabled": 0})
		if scoped:
			return scoped

	return _first_existing("Warehouse", {"is_group": 0, "disabled": 0})


def _resolve_price_list(preferred, buying_or_selling):
	"""Configured price list, else an enabled one of the right type.

	Falls back to creating nothing — a missing price list is non-fatal, the
	caller simply skips writing Item Price rows.
	"""
	if preferred and frappe.db.exists("Price List", preferred):
		return preferred

	filters = {"enabled": 1, buying_or_selling: 1}
	return _first_existing("Price List", filters)


def _import_config(for_write=True, warehouse=None):
	"""Resolve every configurable default, auto-discovering what is not set.

	Nothing is hardcoded: each value comes from ERPNext settings when present,
	otherwise from whatever valid record actually exists on the site. Only a site
	with genuinely no Company / Item Group / UOM / Warehouse fails, because those
	four are mandatory on Item and Stock Reconciliation and cannot be invented.

	`for_write=False` is used by the read-only list/export paths: they only need
	the values to look prices and balances up with, so an incomplete site should
	yield empty columns rather than an error.

	`warehouse` overrides the auto-discovered default with an explicit choice from
	the import screen. It is validated against the resolved company so a caller
	cannot write stock into another company's warehouse.
	"""
	# The inventory module must work on a site that has no POS Profile yet, so a
	# missing/incomplete POS config degrades to auto-discovery instead of throwing.
	try:
		config = resolve_config()
	except Exception:
		config = {}

	stock_settings = frappe.get_cached_doc("Stock Settings")
	buying_settings = frappe.get_cached_doc("Buying Settings")

	company = _resolve_company(config.get("company"))
	item_group = _resolve_item_group(stock_settings.item_group)
	stock_uom = _resolve_stock_uom(stock_settings.stock_uom)

	chosen = (warehouse or "").strip()
	if chosen:
		# Never trust a client-supplied warehouse: it must exist, be a real
		# (non-group) stock location, and belong to the resolved company.
		details = frappe.db.get_value(
			"Warehouse", chosen, ["name", "company", "is_group"], as_dict=True
		)
		if not details:
			frappe.throw(_("Warehouse {0} does not exist.").format(chosen))
		if details.is_group:
			frappe.throw(
				_("{0} is a group warehouse. Choose a warehouse that holds stock.").format(chosen)
			)
		if company and details.company != company:
			frappe.throw(
				_("Warehouse {0} belongs to {1}, not {2}.").format(
					chosen, details.company, company
				)
			)
		warehouse = details.name
	else:
		warehouse = _resolve_warehouse(
			company, stock_settings.default_warehouse or config.get("warehouse")
		)

	# Only genuinely unsatisfiable state is an error, and only when we are about to
	# write. Each of these is mandatory on the documents we create, so there is no
	# safe fallback left to try.
	if for_write:
		missing = []
		if not company:
			missing.append(_("a Company"))
		if not item_group:
			missing.append(_("an Item Group"))
		if not stock_uom:
			missing.append(_("a Unit of Measure"))
		if not warehouse:
			missing.append(_("a Warehouse (non-group)"))
		if missing:
			frappe.throw(
				_("This site has no {0}. Create one in ERPNext, then retry the import.").format(
					", ".join(missing)
				)
			)

	return {
		"company": company,
		"warehouse": warehouse,
		"item_group": item_group,
		"stock_uom": stock_uom,
		"selling_price_list": _resolve_price_list(config.get("price_list"), "selling"),
		"buying_price_list": _resolve_price_list(buying_settings.buying_price_list, "buying"),
		"supplier_group": buying_settings.supplier_group or _first_existing("Supplier Group", {"is_group": 0}),
	}


def _generate_barcode():
	"""Generate a unique 12-digit numeric barcode.

	Uses the cryptographically-seeded stdlib RNG and re-rolls on collision. The
	UNIQUE constraint on `Item Barcode.barcode` is the ultimate guard; this loop
	just avoids the round-trip failure in the common case.
	"""
	import random

	for _attempt in range(50):
		# First digit non-zero so the value is always exactly 12 characters.
		candidate = str(random.randint(100000000000, 999999999999))
		if not frappe.db.exists("Item Barcode", {"barcode": candidate}):
			return candidate

	frappe.throw(_("Could not allocate a unique barcode. Please retry."))


def _existing_barcode(item_code):
	"""Return the item's current barcode, or None. Barcodes never change."""
	return frappe.db.get_value("Item Barcode", {"parent": item_code}, "barcode")


def _ensure_barcode(doc):
	"""Guarantee exactly one barcode on the Item document. Idempotent.

	Operates on the in-memory doc so the caller controls when it is saved.
	"""
	if doc.get("barcodes"):
		return doc.barcodes[0].barcode

	barcode = _generate_barcode()
	# barcode_type intentionally left empty — see Section 10 design notes.
	doc.append("barcodes", {"barcode": barcode})
	return barcode


def _ensure_supplier(supplier_name, config):
	"""Find or create a Supplier by name. Never creates duplicates.

	Matched on the normalized key rather than raw equality. An Arabic supplier
	name carrying NBSP or bidi marks does not compare equal to the same name
	already stored, so the lookup missed, the insert then collided on the unique
	supplier name, and the row failed with the supplier left unlinked.
	"""
	supplier_name = _normalize_text(supplier_name)
	if not supplier_name:
		return None

	key = _match_key(supplier_name)

	# Supplier may be named by series, so both the name and the supplier_name
	# field are compared. Narrowed in SQL by the longest plain run, then matched
	# in Python so the stored side is normalized too.
	anchor = max(supplier_name.split(), key=len, default=supplier_name)
	for row in frappe.get_all(
		"Supplier",
		or_filters={
			"supplier_name": ["like", f"%{anchor}%"],
			"name": ["like", f"%{anchor}%"],
		},
		fields=["name", "supplier_name"],
		order_by="creation asc",
		limit=50,
	):
		if _match_key(row.supplier_name) == key or _match_key(row.name) == key:
			return row.name

	doc = frappe.new_doc("Supplier")
	doc.supplier_name = supplier_name
	doc.supplier_type = "Company"  # mandatory on Supplier; matches DocType default
	if config.get("supplier_group"):
		doc.supplier_group = config["supplier_group"]
	# Swift Storekeeper is a custom role with no Supplier permission in ERPNext, so
	# an ordinary insert raised PermissionError and the row failed with the supplier
	# left unlinked. The caller is already gated on that role and every field here is
	# server-controlled, so the write is elevated rather than widening the role's
	# DocType permissions (which would also open the desk UI and /api/resource).
	doc.insert(ignore_permissions=True)
	return doc.name


def _set_item_price(item_code, price_list, rate, uom):
	"""Create or update the Item Price for a price list. Idempotent.

	Item Price.check_duplicates() forbids a second row for the same
	item/price-list/UOM, so an existing row must be updated in place.
	"""
	if not price_list or rate is None:
		return

	rate = flt(rate)
	existing = frappe.db.get_value(
		"Item Price",
		{"item_code": item_code, "price_list": price_list, "uom": uom},
		"name",
	)

	if existing:
		if flt(frappe.db.get_value("Item Price", existing, "price_list_rate")) != rate:
			doc = frappe.get_doc("Item Price", existing)
			doc.price_list_rate = rate
			doc.save(ignore_permissions=True)
		return

	# Skip creating a zero-rate row; it carries no information.
	if rate <= 0:
		return

	doc = frappe.new_doc("Item Price")
	doc.item_code = item_code
	doc.price_list = price_list
	doc.uom = uom
	doc.price_list_rate = rate
	# buying/selling/currency are derived from the Price List by the controller.
	# Elevated: every caller is gated on Swift Storekeeper, and item/price list/uom
	# are resolved server-side, so the client cannot steer which row is written.
	doc.insert(ignore_permissions=True)


def _find_item_by_name(item_name):
	"""Resolve an Item by name, comparing on the normalized key.

	Returns (item_code, error). item_name is NOT unique in ERPNext, so an
	ambiguous match is reported rather than silently overwriting one of them.

	The comparison cannot be a plain equality filter. Items imported before the
	sheet was normalized are stored WITH their NBSP and bidi marks, so an exact
	filter misses them, the row is treated as new, and the insert then collides on
	Item.name and the row disappears from the import. Candidates are narrowed in
	SQL by the longest run of plain characters in the name, then compared in
	Python on _match_key so both the stored and the incoming form are cleaned.
	"""
	normalized = _normalize_text(item_name)
	if not normalized:
		return None, None

	key = _match_key(normalized)

	# Narrow in SQL first so this never scans the whole Item table. The longest
	# space-free run is used because it is the part least likely to contain the
	# invisible characters being normalized away.
	anchor = max(normalized.split(), key=len, default=normalized)
	candidates = frappe.get_all(
		"Item",
		or_filters={
			"item_name": ["like", f"%{anchor}%"],
			"name": ["like", f"%{anchor}%"],
		},
		fields=["name", "item_name"],
		order_by="creation asc",
		limit=50,
	)

	matches = [c for c in candidates if _match_key(c.item_name) == key or _match_key(c.name) == key]

	if not matches:
		return None, None
	if len(matches) > 1:
		names = ", ".join(m.name for m in matches)
		return None, _(
			"Ambiguous: {0} items already share this name ({1}). Resolve manually."
		).format(len(matches), names)

	return matches[0].name, None


def _parse_import_rows(file_content):
	"""Read an .xlsx payload into normalized row dicts.

	Returns (rows, columns). Raises a clear error on an unreadable file or
	missing required columns.
	"""
	from io import BytesIO

	try:
		from openpyxl import load_workbook

		workbook = load_workbook(
			filename=BytesIO(file_content), read_only=True, data_only=True
		)
	except Exception:
		frappe.throw(
			_("Could not read the file. Please upload a valid .xlsx spreadsheet.")
		)

	try:
		sheet = workbook.active
		raw_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
	finally:
		workbook.close()

	if not raw_rows:
		frappe.throw(_("The spreadsheet is empty."))

	# Locate the header row: the first row that maps to a known column.
	header_index, mapping = None, {}
	for index, row in enumerate(raw_rows[:10]):
		candidate = {}
		for position, cell in enumerate(row):
			key = _match_key(cell)
			if key in IMPORT_COLUMN_ALIASES:
				candidate[IMPORT_COLUMN_ALIASES[key]] = position
		if "item_name" in candidate:
			header_index, mapping = index, candidate
			break

	if header_index is None:
		frappe.throw(
			_("Could not find a header row. Expected columns such as: {0}.").format(
				"Name, QTY, Supplier, Cost Price, Selling Price"
			)
		)

	missing = [c for c in REQUIRED_IMPORT_COLUMNS if c not in mapping]
	if missing:
		labels = {"item_name": _("Name"), "qty": _("QTY")}
		frappe.throw(
			_("Missing required column(s): {0}.").format(
				", ".join(str(labels.get(m, m)) for m in missing)
			)
		)

	def cell(row, field):
		position = mapping.get(field)
		if position is None or position >= len(row):
			return ""
		# _normalize_text, not .strip(): Excel pads Arabic cells with NBSP and bidi
		# marks that .strip() leaves in place, and those survive into the item name,
		# the dedupe key and the Supplier lookup, where they silently split one item
		# into two and stop a supplier from ever matching.
		return _normalize_text(row[position])

	rows = []
	for offset, raw in enumerate(raw_rows[header_index + 1 :]):
		# Ignore completely empty rows.
		if not raw or all(c is None or str(c).strip() == "" for c in raw):
			continue

		rows.append(
			{
				"row_number": header_index + offset + 2,  # 1-based, matches Excel
				"item_name": cell(raw, "item_name"),
				"qty": cell(raw, "qty"),
				"supplier": cell(raw, "supplier"),
				"cost_price": cell(raw, "cost_price"),
				"selling_price": cell(raw, "selling_price"),
			}
		)

	return rows, sorted(mapping.keys())


def _validate_import_row(row):
	"""Validate one parsed row. Returns (clean_dict, error_message)."""
	item_name = row["item_name"]
	if not item_name:
		return None, _("Name is required.")
	if len(item_name) > 140:
		return None, _("Name exceeds 140 characters.")

	def number(raw, label, default=None, required=False):
		if raw in ("", None):
			if required:
				return None, _("{0} is required.").format(label)
			return default, None
		try:
			value = flt(raw)
		except Exception:
			return None, _("{0} is not a valid number: {1}").format(label, raw)
		if value < 0:
			return None, _("{0} cannot be negative.").format(label)
		return value, None

	qty, error = number(row["qty"], _("QTY"), required=True)
	if error:
		return None, error

	cost_price, error = number(row["cost_price"], _("Cost Price"), default=None)
	if error:
		return None, error

	selling_price, error = number(row["selling_price"], _("Selling Price"), default=0)
	if error:
		return None, error

	return {
		"row_number": row["row_number"],
		"item_name": item_name,
		"qty": qty,
		"supplier": row["supplier"],
		"cost_price": cost_price,
		"selling_price": selling_price,
		# Which cells the sheet actually filled in. Needed when merging duplicate
		# rows: a blank price must not overwrite a real one, and blank/0 are
		# indistinguishable once the defaults above have been applied.
		"provided": {
			"supplier": bool(row["supplier"]),
			"cost_price": row["cost_price"] not in ("", None),
			"selling_price": row["selling_price"] not in ("", None),
		},
	}, None


def _collapse_duplicate_rows(rows):
	"""Merge rows that name the same item, summing their quantities.

	A sheet legitimately lists one item on several lines (separate deliveries,
	pack sizes, one line per crate). Those are additive, so the quantities are
	summed rather than one line winning.

	For the non-additive fields the LAST row that actually fills the cell wins,
	matching the "imported Excel is the latest source of truth" rule. A blank
	cell never overwrites a value an earlier row supplied.

	Matching is case-insensitive on the trimmed item name, which is the same key
	the sheet is reconciled against elsewhere. Returns the merged rows in the
	order each item first appeared.
	"""
	merged = {}

	for clean in rows:
		key = _match_key(clean["item_name"])
		existing = merged.get(key)

		if existing is None:
			clean["merged_from"] = [clean["row_number"]]
			merged[key] = clean
			continue

		existing["qty"] = flt(existing["qty"]) + flt(clean["qty"])

		# Later rows only override where they actually carry a value.
		for field in ("supplier", "cost_price", "selling_price"):
			if clean["provided"][field]:
				existing[field] = clean[field]
				existing["provided"][field] = True

		existing["merged_from"].append(clean["row_number"])

	return list(merged.values())


def _merge_import_rows(rows):
	"""Validate then collapse raw parsed rows. Returns (merged_rows, errors).

	Both the preview and the commit go through here, so the two can never disagree
	about how a sheet is interpreted.
	"""
	parsed, errors = [], []

	for row in rows:
		clean, error = _validate_import_row(row)
		if error:
			errors.append(
				{
					"row": row["row_number"],
					"item_name": row["item_name"],
					"supplier": row.get("supplier") or "",
					"qty": row.get("qty"),
					"warehouse": "",
					"exception": "ValidationError",
					"error": error,
					"traceback": "",
				}
			)
			continue
		parsed.append(clean)

	return _collapse_duplicate_rows(parsed), errors


def _read_uploaded_xlsx():
	"""Return the bytes of the uploaded spreadsheet from the current request."""
	files = getattr(frappe.request, "files", None)
	file_object = files.get("file") if files else None
	if not file_object:
		frappe.throw(_("No file was uploaded."))

	filename = (file_object.filename or "").lower()
	if not filename.endswith(".xlsx"):
		frappe.throw(
			_("Only .xlsx files are supported. Please re-save the sheet as .xlsx.")
		)

	content = file_object.stream.read()
	if not content:
		frappe.throw(_("The uploaded file is empty."))

	max_bytes = 10 * 1024 * 1024
	if len(content) > max_bytes:
		frappe.throw(_("File is too large. Maximum size is 10 MB."))

	return content


@frappe.whitelist(methods=["POST"])
def inventory_import_preview():
	"""Parse + validate an uploaded sheet WITHOUT writing anything.

	Lets the Storekeeper see exactly what will be created vs updated, and which
	rows will be skipped, before committing.
	"""
	require_role("Swift Storekeeper")
	config = _import_config(warehouse=frappe.form_dict.get("warehouse"))

	rows, columns = _parse_import_rows(_read_uploaded_xlsx())

	# Repeated items are summed into one line before anything is looked up, so the
	# preview shows the same totals the commit will write.
	valid, errors = _merge_import_rows(rows)

	resolved = []
	for clean in valid:
		item_code, ambiguity = _find_item_by_name(clean["item_name"])
		if ambiguity:
			errors.append(
				{"row": clean["row_number"], "item_name": clean["item_name"], "error": ambiguity}
			)
			continue

		clean["action"] = "update" if item_code else "create"
		clean["item_code"] = item_code
		clean["supplier_exists"] = bool(
			clean["supplier"]
			and frappe.db.get_value("Supplier", {"supplier_name": clean["supplier"]}, "name")
		)
		resolved.append(clean)

	valid = resolved

	new_suppliers = sorted(
		{r["supplier"] for r in valid if r["supplier"] and not r["supplier_exists"]}
	)

	return {
		"columns_detected": columns,
		"config": config,
		"total_rows": len(rows),
		"valid_count": len(valid),
		"create_count": sum(1 for r in valid if r["action"] == "create"),
		"update_count": sum(1 for r in valid if r["action"] == "update"),
		# How many sheet lines were folded away by merging, so the Storekeeper can
		# reconcile "50 rows in my file" against a smaller item count.
		"merged_count": sum(len(r["merged_from"]) - 1 for r in valid),
		"new_suppliers": new_suppliers,
		"rows": valid,
		"errors": errors,
	}


def _apply_import_row(clean, config):
	"""Create or update a single Item from a validated row.

	Returns a result dict. Runs inside the caller's savepoint.
	"""
	supplier = _ensure_supplier(clean["supplier"], config) if clean["supplier"] else None
	item_code, ambiguity = _find_item_by_name(clean["item_name"])
	if ambiguity:
		frappe.throw(ambiguity)

	frappe.logger("swift_import").debug(
		f"apply row={clean['row_number']} item={clean['item_name']!r} "
		f"qty={clean['qty']} supplier={supplier!r} warehouse={config['warehouse']} "
		f"action={'create' if item_code is None else 'update'}"
	)

	created = item_code is None

	if created:
		doc = frappe.new_doc("Item")
		doc.item_name = clean["item_name"]
		doc.item_group = config["item_group"]
		doc.stock_uom = config["stock_uom"]
		doc.is_stock_item = 1

		naming_by = (
			frappe.db.get_single_value("Stock Settings", "item_naming_by") or "Item Code"
		)
		if naming_by == "Item Code":
			# Item.name is derived from item_code, and item_name is NOT unique, so two
			# different sheet rows can want the same Item.name -- most often Arabic
			# names that differ only by characters ERPNext strips while naming.
			# _find_item_by_name already proved no Item carries this item_name, so a
			# record sitting under this code is a DIFFERENT item: the insert would
			# raise DuplicateEntryError, the row would be recorded as failed, and the
			# item would never be created even though the preview promised "Create".
			# Suffixing keeps the row importable and leaves item_name untouched.
			code = clean["item_name"]
			if frappe.db.exists("Item", code):
				suffix = 2
				while frappe.db.exists("Item", f"{code[:135]}-{suffix}"):
					suffix += 1
				code = f"{code[:135]}-{suffix}"
			doc.item_code = code

		if clean["cost_price"] is not None:
			doc.valuation_rate = flt(clean["cost_price"])

		defaults = {
			"company": config["company"],
			"default_warehouse": config["warehouse"],
		}
		if supplier:
			defaults["default_supplier"] = supplier
		doc.append("item_defaults", defaults)

		# The Item Supplier grid is a SEPARATE child table from Item Default. Setting
		# default_supplier alone leaves supplier_items empty, which is why suppliers
		# looked unlinked on the item even though the import reported success.
		if supplier:
			doc.append("supplier_items", {"supplier": supplier})

		barcode = _ensure_barcode(doc)
		# Elevated for the same reason as the Supplier insert: the role is checked at
		# the endpoint and every field set above comes from validated, server-side
		# values. This is why "Preview says Create" but the Item never appeared.
		doc.insert(ignore_permissions=True)
		item_code = doc.name
	else:
		doc = frappe.get_doc("Item", item_code)

		if clean["cost_price"] is not None:
			doc.valuation_rate = flt(clean["cost_price"])

		# The defaults row is realigned on EVERY update, not only when the sheet
		# names a supplier. Previously this whole block was skipped for a row with
		# no supplier, and the warehouse was only filled in when it was empty, so an
		# item imported earlier into another warehouse kept pointing at that one
		# while its stock was written to the chosen warehouse.
		row = next((d for d in doc.item_defaults if d.company == config["company"]), None)
		if row:
			row.default_warehouse = config["warehouse"]
			if supplier:
				row.default_supplier = supplier
		else:
			defaults = {
				"company": config["company"],
				"default_warehouse": config["warehouse"],
			}
			if supplier:
				defaults["default_supplier"] = supplier
			doc.append("item_defaults", defaults)

		# Same separate-child-table gap as the create branch. Appended only when the
		# supplier is not already in the grid, so re-importing the same sheet does not
		# stack duplicate rows.
		if supplier and not any(d.supplier == supplier for d in doc.supplier_items):
			doc.append("supplier_items", {"supplier": supplier})

		# Existing barcode is preserved; only generated when absent.
		barcode = _ensure_barcode(doc)
		# Elevated for the same reason as the create branch above.
		doc.save(ignore_permissions=True)

	_set_item_price(
		item_code, config["selling_price_list"], clean["selling_price"], config["stock_uom"]
	)
	if clean["cost_price"] is not None:
		_set_item_price(
			item_code, config["buying_price_list"], clean["cost_price"], config["stock_uom"]
		)

	return {
		"row": clean["row_number"],
		"item_code": item_code,
		"item_name": clean["item_name"],
		"barcode": barcode,
		"supplier": supplier,
		"action": "created" if created else "updated",
		"qty": clean["qty"],
		"warehouse": config["warehouse"],
	}


def _reconcile_stock(applied, config):
	"""Set on-hand qty to the sheet's absolute value via Stock Reconciliation.

	Returns (reconciliation_name, message). ERPNext drops rows with no change and
	raises EmptyStockReconciliationItemsError when nothing differs, which is a
	successful no-op here.

	Each item gets its OWN reconciliation document. A single shared document is
	all-or-nothing: one item ERPNext refuses (most often a new item with no
	valuation rate, because Cost Price is optional in the sheet) aborts the whole
	submit and every other item silently keeps a zero balance. Per-item documents
	mean a rejected item is reported by name and the rest still get their stock.
	"""
	from erpnext.stock.doctype.stock_reconciliation.stock_reconciliation import (
		EmptyStockReconciliationItemsError,
	)

	targets = [r for r in applied if r["qty"] is not None]
	if not targets:
		return None, None

	warehouse = config["warehouse"]
	created, unchanged, failures = [], 0, []

	for result in targets:
		item_code = result["item_code"]
		qty = flt(result["qty"])

		# A valuation rate is mandatory whenever stock is being introduced, and
		# Cost Price is optional in the sheet. Fall back to the last known rate for
		# the item so a priceless row still gets its quantity instead of aborting.
		rate = flt(frappe.db.get_value("Item", item_code, "valuation_rate"))
		if rate <= 0:
			rate = flt(
				frappe.db.get_value(
					"Bin",
					{"item_code": item_code, "warehouse": warehouse},
					"valuation_rate",
				)
			)

		row = {"item_code": item_code, "warehouse": warehouse, "qty": qty}
		if rate > 0:
			row["valuation_rate"] = rate
		else:
			# No rate is discoverable anywhere and Cost Price is optional in the sheet.
			# valuation_rate = 0.0 alone is NOT enough: ERPNext's validator tests the
			# rate for truthiness, so 0 reads as "absent" and it still throws
			# "Valuation Rate required for Item". allow_zero_valuation_rate is the
			# field that makes a zero rate legitimate, and it is set for BOTH new and
			# existing items -- an item already on file with a 0 rate hits the same
			# wall as a brand new one.
			row["valuation_rate"] = 0.0
			row["allow_zero_valuation_rate"] = 1
			# Set only when the installed ERPNext actually carries the field. Assigning
			# an unknown fieldname on a child row is silently kept as a stray attribute
			# on some versions and rejected on others, so it is probed rather than
			# assumed.
			if frappe.get_meta("Stock Reconciliation Item").has_field(
				"valuation_rate_per_unit"
			):
				row["valuation_rate_per_unit"] = 0.0

		frappe.logger("swift_import").debug(
			f"stock reco: item={item_code} warehouse={warehouse} qty={qty} rate={rate}"
		)

		savepoint = f"reco_{len(created) + len(failures)}"
		try:
			frappe.db.savepoint(savepoint)
			doc = frappe.new_doc("Stock Reconciliation")
			doc.company = config["company"]
			doc.purpose = "Stock Reconciliation"
			# expense_account and cost_center are resolved from Company defaults by
			# the controller's validate().
			doc.append("items", row)
			# Elevated: Swift Storekeeper has no Stock Reconciliation permission, so
			# every row failed here and the item landed with Qty = 0. The flag is set
			# on the document because submit() re-checks permissions independently of
			# insert(). Company, warehouse and qty are all server-validated.
			doc.flags.ignore_permissions = True
			doc.insert(ignore_permissions=True)
			doc.submit()
			created.append(doc.name)
		except EmptyStockReconciliationItemsError:
			# Already at the requested qty and rate; nothing to adjust.
			unchanged += 1
		except Exception as exc:
			frappe.db.rollback(save_point=savepoint)
			# Exception class included: DuplicateEntryError and friends have an empty
			# str(), which previously collapsed to a blank, unactionable message.
			detail = frappe.utils.strip_html(str(exc)) or type(exc).__name__
			failures.append(f"{result['item_name']}: {detail}")
			frappe.log_error(
				title=f"Import stock failed: {result['item_name']}",
				message=(
					f"item_code={item_code} warehouse={warehouse} qty={qty} rate={rate}\n"
					f"{type(exc).__name__}: {exc}\n\n{frappe.get_traceback()}"
				),
			)

	message = None
	if failures:
		shown = "; ".join(failures[:5])
		if len(failures) > 5:
			shown += _(" (+{0} more)").format(len(failures) - 5)
		message = _("Stock could not be set for {0} item(s) in {1}: {2}").format(
			len(failures), warehouse, shown
		)
	elif not created and unchanged:
		message = _("Stock quantities already matched the sheet; no adjustment needed.")

	return (created[0] if len(created) == 1 else created or None), message


@frappe.whitelist(methods=["POST"])
def inventory_import_commit():
	"""Import the uploaded sheet: suppliers, items, barcodes, prices and stock.

	Each row is isolated in its own savepoint, so one bad row cannot roll back
	the rest of the sheet.
	"""
	require_role("Swift Storekeeper")
	config = _import_config(warehouse=frappe.form_dict.get("warehouse"))

	rows, _columns = _parse_import_rows(_read_uploaded_xlsx())
	if not rows:
		frappe.throw(_("The spreadsheet contains no data rows."))

	applied, errors = [], []

	# Same collapse the preview showed, so what gets written matches what was shown.
	merged, merge_errors = _merge_import_rows(rows)
	errors.extend(merge_errors)

	for clean in merged:
		savepoint = f"row_{clean['row_number']}"
		try:
			frappe.db.savepoint(savepoint)
			applied.append(_apply_import_row(clean, config))
		except Exception as exc:
			frappe.db.rollback(save_point=savepoint)
			# Every failed row reports the full picture. str(exc) alone is empty for
			# DuplicateEntryError and several other Frappe exceptions, which is how
			# rows used to disappear behind "Could not import this row."
			traceback = frappe.get_traceback()
			errors.append(
				{
					"row": clean["row_number"],
					"item_name": clean["item_name"],
					"supplier": clean.get("supplier") or "",
					"qty": clean.get("qty"),
					"warehouse": config["warehouse"],
					"exception": type(exc).__name__,
					"error": (
						frappe.utils.strip_html(str(exc))
						or f"{type(exc).__name__} (no message)"
					),
					"traceback": traceback,
				}
			)
			frappe.log_error(
				title=f"Import row {clean['row_number']} failed: {clean['item_name']}",
				message=(
					f"row={clean['row_number']} item={clean['item_name']!r} "
					f"qty={clean.get('qty')} supplier={clean.get('supplier')!r} "
					f"warehouse={config['warehouse']}\n"
					f"{type(exc).__name__}: {exc}\n\n{traceback}"
				),
			)

	stock_entry, stock_message = None, None
	if applied:
		savepoint = "stock_reco"
		try:
			frappe.db.savepoint(savepoint)
			stock_entry, stock_message = _reconcile_stock(applied, config)
		except Exception as exc:
			frappe.db.rollback(save_point=savepoint)
			# Items are still imported; surface the stock failure without losing them.
			stock_message = _("Items imported, but stock could not be updated: {0}").format(
				frappe.utils.strip_html(str(exc))
			)

	applied_rows = {r["row"] for r in applied}

	return {
		"success": True,
		"total_rows": len(rows),
		"imported": len(applied),
		"created": sum(1 for r in applied if r["action"] == "created"),
		"updated": sum(1 for r in applied if r["action"] == "updated"),
		# Sheet lines folded away by merging. Explains why `imported` can be lower
		# than `total_rows` without any row having been skipped. Counted only for
		# rows that actually got written, so a merge inside a failed row is not
		# reported as if its quantities had been applied.
		"merged": sum(
			len(r["merged_from"]) - 1 for r in merged if r["row_number"] in applied_rows
		),
		"skipped": len(errors),
		# Echo where the stock actually went, so the Storekeeper can confirm the
		# import landed in the warehouse they picked.
		"warehouse": config["warehouse"],
		"stock_reconciliation": stock_entry,
		"stock_message": stock_message,
		"results": applied,
		"errors": errors,
	}


def _inventory_rows(search=None, supplier=None, barcode=None, limit=100, start=0):
	"""Shared query behind the inventory screen and the Excel export."""
	# Read-only: never block listing/exporting on a write prerequisite.
	config = _import_config(for_write=False)

	filters = {"disabled": 0}

	# Intersect the name-based filters so combining them narrows the result.
	restrict_to = None

	if barcode:
		# Partial match so a half-typed / partially-scanned barcode still finds rows.
		owners = set(
			frappe.get_all(
				"Item Barcode",
				filters={"barcode": ["like", f"%{barcode.strip()}%"], "parenttype": "Item"},
				pluck="parent",
			)
		)
		if not owners:
			return []
		restrict_to = owners

	if supplier:
		supplier_items = set(
			frappe.get_all(
				"Item Default",
				filters={"default_supplier": supplier, "parenttype": "Item"},
				pluck="parent",
			)
		)
		if not supplier_items:
			return []
		restrict_to = supplier_items if restrict_to is None else (restrict_to & supplier_items)

	if search:
		# Partial, case-insensitive match across every identifier the storekeeper
		# might type: item name, item code, barcode, or supplier name. LIKE in
		# MariaDB is case-insensitive under the default collation.
		raw = search.strip()
		term = f"%{raw}%"

		matches = set(
			frappe.get_all(
				"Item",
				filters={"item_name": ["like", term]},
				pluck="name",
			)
		)
		matches.update(
			frappe.get_all(
				"Item",
				filters={"name": ["like", term]},
				pluck="name",
			)
		)
		matches.update(
			frappe.get_all(
				"Item Barcode",
				filters={"barcode": ["like", term], "parenttype": "Item"},
				pluck="parent",
			)
		)

		# Suppliers are frequently named by naming series, so resolve the typed
		# text against supplier_name and then map back to items.
		supplier_names = frappe.get_all(
			"Supplier",
			filters={"supplier_name": ["like", term]},
			pluck="name",
		)
		if supplier_names:
			matches.update(
				frappe.get_all(
					"Item Default",
					filters={"default_supplier": ["in", supplier_names], "parenttype": "Item"},
					pluck="parent",
				)
			)

		if not matches:
			return []

		# Intersect rather than replace, so search combines with the other filters.
		restrict_to = matches if restrict_to is None else (restrict_to & matches)
		if not restrict_to:
			return []

	if restrict_to is not None:
		if not restrict_to:
			return []
		filters["name"] = ["in", list(restrict_to)]

	items = frappe.get_all(
		"Item",
		filters=filters,
		fields=["name as item_code", "item_name", "stock_uom", "valuation_rate", "item_group"],
		order_by="modified desc",
		limit=cint(limit) or 100,
		start=cint(start),
	)
	if not items:
		return []

	codes = [i.item_code for i in items]

	barcodes = {}
	for entry in frappe.get_all(
		"Item Barcode",
		filters={"parent": ["in", codes]},
		fields=["parent", "barcode"],
	):
		barcodes.setdefault(entry.parent, entry.barcode)

	suppliers = {}
	for entry in frappe.get_all(
		"Item Default",
		filters={"parent": ["in", codes], "parenttype": "Item"},
		fields=["parent", "default_supplier"],
	):
		if entry.default_supplier:
			suppliers.setdefault(entry.parent, entry.default_supplier)

	def price_map(price_list):
		if not price_list:
			return {}
		result = {}
		for entry in frappe.get_all(
			"Item Price",
			filters={"item_code": ["in", codes], "price_list": price_list},
			fields=["item_code", "price_list_rate"],
		):
			result.setdefault(entry.item_code, flt(entry.price_list_rate))
		return result

	selling = price_map(config["selling_price_list"])
	buying = price_map(config["buying_price_list"])

	balances = {}
	if config.get("warehouse"):
		for entry in frappe.get_all(
			"Bin",
			filters={"item_code": ["in", codes], "warehouse": config["warehouse"]},
			fields=["item_code", "actual_qty"],
		):
			balances[entry.item_code] = flt(entry.actual_qty)

	for item in items:
		item["barcode"] = barcodes.get(item.item_code)
		item["supplier"] = suppliers.get(item.item_code)
		item["selling_price"] = selling.get(item.item_code, 0)
		item["cost_price"] = buying.get(item.item_code, flt(item.get("valuation_rate")) or 0)
		item["qty"] = balances.get(item.item_code, 0)

	return items


@frappe.whitelist(methods=["GET"])
def inventory_list(search=None, supplier=None, barcode=None, limit=100, start=0):
	"""Paged inventory for the Storekeeper screen."""
	require_role("Swift Storekeeper")
	return _inventory_rows(search, supplier, barcode, limit, start)


@frappe.whitelist(methods=["GET"])
def inventory_export(search=None, supplier=None, barcode=None):
	"""Stream the inventory back as .xlsx, mirroring the import layout.

	Newly created items appear automatically because this reads live data rather
	than a stored snapshot.
	"""
	require_role("Swift Storekeeper")

	from frappe.utils.xlsxutils import make_xlsx

	rows = _inventory_rows(search, supplier, barcode, limit=10000, start=0)

	data = [[label for _key, label in EXPORT_COLUMNS]]
	for row in rows:
		data.append(
			[
				row.get("item_name") or "",
				# Keep the barcode textual so Excel cannot render 12 digits in
				# scientific notation and corrupt the round-trip.
				str(row.get("barcode") or ""),
				flt(row.get("qty")),
				row.get("supplier") or "",
				flt(row.get("cost_price")),
				flt(row.get("selling_price")),
			]
		)

	xlsx_file = make_xlsx(data, "Inventory")

	# Frappe's `as_binary` responder reads exactly these three keys and streams the
	# bytes as application/octet-stream. Anything that raises before this point is
	# rendered by `as_json` instead, and a bare frappe.throw is a ValidationError,
	# which carries http_status_code = 417 — that was the reported export failure.
	frappe.local.response.filename = "inventory.xlsx"
	frappe.local.response.filecontent = xlsx_file.getvalue()
	frappe.local.response.type = "binary"


@frappe.whitelist(methods=["PUT"])
def update_inventory_item(
	item_code=None,
	item_name=None,
	supplier=None,
	cost_price=None,
	selling_price=None,
	barcode=None,
	qty=None,
):
	"""Edit one item from the Storekeeper screen.

	Only the fields the business allows are touched — this is deliberately an
	explicit signature rather than **fields, so no caller can reach arbitrary
	Item columns such as valuation_rate or disabled.
	"""
	require_role("Swift Storekeeper")

	if not item_code:
		frappe.throw(_("item_code is required."))
	if not frappe.db.exists("Item", item_code):
		frappe.local.response.http_status_code = 404
		frappe.throw(_("Item {0} not found.").format(item_code))

	config = _import_config()
	doc = frappe.get_doc("Item", item_code)

	if item_name is not None:
		item_name = str(item_name).strip()
		if not item_name:
			frappe.throw(_("Item name cannot be empty."))
		if item_name != doc.item_name:
			clash, _ambiguity = _find_item_by_name(item_name)
			if clash and clash != item_code:
				frappe.local.response.http_status_code = 409
				frappe.throw(_("Another item already uses the name {0}.").format(item_name))
			doc.item_name = item_name

	for label, value in ((_("Cost Price"), cost_price), (_("Selling Price"), selling_price)):
		if value not in (None, "") and flt(value) < 0:
			frappe.throw(_("{0} cannot be negative.").format(label))

	if cost_price not in (None, ""):
		doc.valuation_rate = flt(cost_price)

	if supplier is not None:
		supplier = str(supplier).strip()
		resolved = _ensure_supplier(supplier, config) if supplier else None
		row = next((d for d in doc.item_defaults if d.company == config["company"]), None)
		if row:
			row.default_supplier = resolved
			if not row.default_warehouse:
				row.default_warehouse = config["warehouse"]
		elif resolved:
			doc.append(
				"item_defaults",
				{
					"company": config["company"],
					"default_warehouse": config["warehouse"],
					"default_supplier": resolved,
				},
			)

	if barcode is not None:
		barcode = str(barcode).strip()
		if barcode:
			if not barcode.isdigit() or len(barcode) != 12:
				frappe.throw(_("Barcode must be exactly 12 digits."))
			owner = _barcode_owner(barcode)
			if owner and owner != item_code:
				frappe.local.response.http_status_code = 409
				frappe.throw(_("Barcode already assigned to item {0}.").format(owner))
			if doc.barcodes:
				doc.barcodes[0].barcode = barcode
			else:
				doc.append("barcodes", {"barcode": barcode})

	# Every item must always carry exactly one barcode.
	_ensure_barcode(doc)
	# Elevated: gated on Swift Storekeeper, and the signature is explicit so no
	# caller can reach Item fields beyond the ones edited above.
	doc.save(ignore_permissions=True)

	if selling_price not in (None, ""):
		_set_item_price(
			item_code, config["selling_price_list"], flt(selling_price), config["stock_uom"]
		)
	if cost_price not in (None, ""):
		_set_item_price(
			item_code, config["buying_price_list"], flt(cost_price), config["stock_uom"]
		)

	stock_message = None
	if qty not in (None, ""):
		if flt(qty) < 0:
			frappe.throw(_("Quantity cannot be negative."))
		_reconciled, stock_message = _reconcile_stock(
			[{"item_code": item_code, "qty": flt(qty)}], config
		)

	return {
		"success": True,
		"item_code": item_code,
		"barcode": _existing_barcode(item_code),
		"stock_message": stock_message,
	}
